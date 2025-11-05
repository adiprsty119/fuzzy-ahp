from flask import Flask, Response, request, render_template, request as flask_request, redirect, url_for, flash, session, jsonify, current_app
from flask_session import Session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from app import create_app, db
from app.models import Users, Participants, Notification, Event, Kuota, Criteria
from flask_mail import Mail, Message
from twilio.rest import Client
from authlib.integrations.flask_client import OAuth
from markupsafe import escape
from datetime import datetime, timedelta
from flask_wtf import CSRFProtect
from flask_wtf.csrf import CSRFProtect, CSRFError
from forms import LoginForm, RegisterForm
from config import Config
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import current_user, LoginManager, login_user, login_required
from functools import wraps
from app.utils.utils import log_activity
from sqlalchemy.exc import IntegrityError
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import random, string
import logging
import secrets
import time
import os
import re
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

app = create_app()
app.config['SESSION_FILE_PATH'] = os.path.join(app.root_path, 'flask_session')
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.secret_key = os.getenv("APP_SECRET_KEY")
Session(app)

csrf = CSRFProtect(app)
app.config.from_object(Config)
limiter = Limiter(get_remote_address, app=app)
logging.basicConfig(filename='login.log', level=logging.INFO,
                    format='%(asctime)s %(levelname)s:%(message)s')

# Inisialisasi Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)

login_manager.login_view = 'login'  # nama fungsi view untuk login
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return Users.query.get(int(user_id))

# Buat folder uploads jika belum ada
# if not os.path.exists(app.config['UPLOAD_FOLDER']):
    # os.makedirs(app.config['UPLOAD_FOLDER'])
    
# Tentukan folder upload (misalnya di static/uploads)
UPLOAD_FOLDER = os.path.join(app.root_path, 'static/uploads')
ALLOWED_EXTENSIONS = {'csv', 'xls', 'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Configure Flask-Mail OTP
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = os.getenv("MAIL_USERNAME")
app.config['MAIL_PASSWORD'] = os.getenv("MAIL_PASSWORD")
mail = Mail(app)

# Whatsapp OTP
account_sid = os.getenv("TWILIO_ACCOUNT_SID")
auth_token = os.getenv("TWILIO_AUTH_TOKEN")
client = Client(account_sid, auth_token)

def send_whatsapp_code(phone, code):
    account_sid = os.getenv("TWILIO_ACCOUNT_SID")
    auth_token = os.getenv("TWILIO_AUTH_TOKEN")
    client = Client(account_sid, auth_token)
    message_body = f"Kode verifikasi Anda adalah: *{code}*.\nJangan bagikan kode ini kepada siapa pun."
    try:
        message = client.messages.create(
            body=message_body,
            from_='whatsapp:+14155238886',
            to=f'whatsapp:{phone}'
        )
        print("Pesan berhasil dikirim:", message.sid)
    except Exception as e:
        print("Gagal mengirim pesan:", e)

def normalize_phone_number(phone):
    phone = phone.strip()
    if phone.startswith('0'):
        return '+62' + phone[1:]
    elif phone.startswith('+62'):
        return phone
    elif phone.startswith('62'):
        return '+' + phone
    else:
        return phone
    
def generate_username(email):
    name_part = email.split('@')[0]
    random_suffix = ''.join(random.choices(string.digits, k=4))
    return f"{name_part}_{random_suffix}"

# Google OAuth Config
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.getenv('GOOGLE_CLIENT_ID'),
    client_secret=os.getenv('GOOGLE_CLIENT_SECRET'),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    access_token_url='https://oauth2.googleapis.com/token',
    access_token_params=None,
    authorize_url='https://accounts.google.com/o/oauth2/auth',
    authorize_params=None,
    api_base_url='https://www.googleapis.com/oauth2/v1/',
    userinfo_endpoint='https://www.googleapis.com/oauth2/v1/userinfo',
    client_kwargs={'scope': 'openid email profile'},
)

# Fungsi untuk mengecek keberadaan username, nomor hp, email dan password serta untuk menghasilkan kode verifikasi 6 digit
def check_username_in_db(username):
    user = Users.query.filter_by(username=username).first()
    return user is not None

def check_phone_in_db(phone):
    user = Users.query.filter_by(nomor_hp=phone).first()
    return user is not None


def check_email_in_db(email):
    user = Users.query.filter_by(email=email).first()
    return user is not None

def check_password_in_db(username, password):
    user = Users.query.filter_by(username=username).first()
    if user:
        return check_password_hash(user.password, password)
    return False

def generate_verification_code():
    return random.randint(100000, 999999)

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    return render_template('csrf_error.html', reason=e.description), 400

@app.errorhandler(429)
def ratelimit_handler(e):
    return render_template("429.html", message="Terlalu banyak percobaan login. Silakan coba lagi nanti."), 429

@app.context_processor
def inject_notifications():
    user = None
    unread_count = 0
    if session.get('username'):
        user = Users.query.filter_by(username=session.get('username')).first()
    elif session.get('user'):
        user = Users.query.filter_by(username=session['user'].get('username')).first()
    if user:
        unread_count = Notification.query.filter_by(user_id=user.id, is_read=False).count()
    return dict(notification_count=unread_count)

# --- Middleware untuk cek login dan role ---
def my_decorator(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # logika tambahan
        return f(*args, **kwargs)
    return decorated_function

# Endpoint login
@app.route('/login/', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        role = request.form.get("role")
        
        # Query user dari database
        user = Users.query.filter_by(username=username).first()
        
        if not user:
            logging.warning(f"Login gagal: username '{username}' tidak ditemukan.")
            flash("Username salah!", "danger")
        elif not check_password_hash(user.password, password):
            logging.warning(f"Login gagal: password salah untuk user '{username}'.")
            flash("Password salah!", "danger")
        if not role or role.lower() != user.level.lower(): 
            # Validasi role dari dropdown dengan role di database
            logging.warning(f"Login gagal: role '{role}' tidak sesuai untuk user '{username}'.")
            flash("Role tidak sesuai!", "danger")
        else:
            login_user(user)
            session['username'] = username  
            session['role'] = user.level
            safe_username = escape(username)
            logging.info(f"User '{username}' berhasil login sebagai {user.level}.")
            flash(f"Login berhasil! Selamat datang, {safe_username}.", "success")
            session['first_time_login'] = True
            
            # Redirect sesuai role
            if user.level == "admin":
                return redirect(url_for('admin_dashboard'))
            elif user.level == "penilai":
                return redirect(url_for('penilai_dashboard'))
            elif user.level == "peserta":
                return redirect(url_for('peserta_dashboard'))
            else:
                # Default jika role tidak dikenali
                return redirect(url_for('login')) 
    return render_template('login.html', form=form)

# Endpoint login with Google
@app.route('/login/google/', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login_google():
    redirect_uri = url_for('login_google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

# Endpoint Callback Login With Google
@app.route('/login/google/callback/')
def login_google_callback():
    try:
        token = google.authorize_access_token()
        resp = google.get('userinfo')  
        resp.raise_for_status()  
        user_info = resp.json()
    except Exception as e:
        logging.warning(f"Login Google gagal: {e}") 
        flash("Gagal login dengan Google. Silakan coba lagi.", "danger")
        return redirect(url_for('login'))
    
    # Proses lanjut jika data user berhasil diambil
    email = user_info.get('email')
    username = user_info.get('name') or "Pengguna"
    picture = user_info.get('picture') or "img/default-user.png"
    
    if not email:
        flash("Email dari akun Google tidak ditemukan.", "danger")
        return redirect(url_for('login'))
    
    # ✅ Validasi hanya email Gmail
    if not email.endswith('@gmail.com'):
        flash("Login hanya diizinkan dengan akun Gmail.", "danger")
        return redirect(url_for('login'))
    
    user = Users.query.filter_by(email=email).first()
    if user:
        # Update foto dari Google jika belum tersimpan
        if not user.foto or user.foto == "img/default-user.png":
            user.foto = user_info.get('picture') or "img/default-user.png"
            db.session.commit()
            
        session['username'] = user.username
        session['user'] = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'nama_lengkap': user.nama_lengkap,
            'foto': user.foto,
            'level': user.level
        }
        session['first_time_login'] = True
        session.modified = True 
        print("✅ Session set:", session.get('user'))
        logging.info(f"User '{user.username}' berhasil login via Google.")
        flash(f"Login berhasil! Selamat datang, {escape(user.nama_lengkap)}.", "success")
        return redirect(url_for('admin_dashboard'))
    else:
        # Jika belum ada, arahkan ke konfirmasi registrasi
        session['pending_user'] = user_info
        logging.warning(f"Percobaan login Google dari email '{email}' belum terdaftar.")
        flash("Akun Google Anda belum terdaftar. Lanjutkan registrasi?", "warning")
        return redirect(url_for('confirm_register'))
    
# Endpoint Fisrt Confirm Register With Google
@app.route('/confirm-register/')
def confirm_register():
    user_info = session.get('pending_user')
    if not user_info:
        flash("Data user tidak ditemukan. Silakan login ulang.", "danger")
        return redirect(url_for('login'))
    form = RegisterForm()
    return render_template("confirm_register.html", user=user_info, form=form)

# Endpoint Confirm Register With Google
@app.route('/confirm-register', methods=['POST'])
def do_register():
    user_info = session.get('pending_user')
    if not user_info:
        flash("Data user tidak ditemukan. Silakan login ulang.", "danger")
        return redirect(url_for('login'))
    
    form = RegisterForm()

    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        email = user_info['email']
        
    # Validasi apakah username atau email sudah digunakan
        if Users.query.filter_by(email=email).first():
            flash("Email sudah digunakan. Silakan login.", "warning")
            return redirect(url_for('login'))
        if Users.query.filter_by(username=username).first():
            flash("Username sudah digunakan.", "warning")
            return redirect(url_for('confirm_register'))
    
        # Simpan ke database
        new_user = Users(
                username=username,
                password=generate_password_hash(secrets.token_urlsafe(12), method='pbkdf2:sha256'),
                nama_lengkap=user_info['name'],
                email=user_info['email'],
                jenis_kelamin=None,
                usia=None,
                foto=user_info['picture'],
                nomor_hp=None,
                level='user',
                reset_token=None,
                token_exp=None
        )
        try:
            db.session.add(new_user)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logging.error(f"Gagal menyimpan user Google baru: {e}")
            flash("Terjadi kesalahan saat registrasi. Coba lagi.", "danger")
            return redirect(url_for('confirm_register'))

        # Set session
        session['user'] = {
            'id': new_user.id,
            'username': new_user.username,
            'email': new_user.email,
            'nama_lengkap': new_user.nama_lengkap,
            'foto': new_user.foto,
            'level': new_user.level
        }
        print("Session setelah login Google:", dict(session))
        logging.info(f"User baru '{username}' berhasil registrasi dan login via Google.")
        flash("Registrasi berhasil! Anda sudah login untuk pertama kali.", "welcome")
        session['username'] = new_user.username
        session['first_time_login'] = True
        return redirect(url_for('index'))
    return render_template("confirm_register.html", user=user_info, form=form)

# Endpoint register
@app.route('/register/', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        full_name = request.form.get('fullName', '').strip()
        email = request.form.get('email', '').strip().lower()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirmPassword', '')
        level = 'peserta'  
        
        # Apakah ada kolom yang kosong?
        if not all([full_name, email, username, password, confirm_password]):
            flash("Semua kolom wajib diisi.", "danger")
            return redirect(url_for('register'))
        
        # Validasi password dengan regex
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, password):
            flash("Password must have at least 8 characters, including uppercase, lowercase, number, and special character.", "danger")
            return redirect(url_for('register'))
        
        # Validasi apakah password dan confirmPassword cocok
        if password != confirm_password:
            flash("Password and Confirm Password must match!", "danger")
            return redirect(url_for('register'))
        
        # Cek keberadaan username dan email di database
        user_exists = check_username_in_db(username)
        email_exists = check_email_in_db(email)
        
        if not user_exists and not email_exists:
           # Enkripsi password
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16) 
            # Masukkan data ke database
            try:
                new_user = Users(
                    username=username,
                    password=hashed_password,
                    nama_lengkap=full_name,
                    email=email,
                    level=level,
                    status='aktif'
                )
                db.session.add(new_user)
                db.session.commit()
                flash("Registrasi berhasil! Selamat datang di sistem kami. Silakan login untuk mulai menggunakan fitur.", "welcome")
                return redirect(url_for('login'))
            except IntegrityError as e:
                db.session.rollback()
                logging.error(f"Error during registration: {e}")
                flash("An error occurred during registration, Please try again.", "danger")
                print(e)
        elif user_exists and email_exists:
            flash("Username dan email Anda telah terdaftar.", "danger")
        elif user_exists:
            flash("Username Anda telah terdaftar.", "danger")
        elif email_exists:
            flash("Email Anda telah terdaftar.", "danger")
        return redirect(url_for('register'))
    return render_template('register.html')

# Endpoint Register With Google
@app.route('/register/google/', methods=['GET', 'POST'])
def register_google():
    redirect_uri = url_for('register_google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

# Endpoint Callback Register With Google
@app.route('/register/google/callback/')
def register_google_callback():
    try:
        token = google.authorize_access_token()
        resp = google.get('userinfo')
        resp.raise_for_status() 
        user_info = resp.json()
    except Exception as e:
        print(f"Google registrasi error: {e}") 
        flash("Gagal melakukan registrasi dengan Google. Silakan coba lagi.", "danger")
        return redirect(url_for('register'))
    
    email = user_info['email']
    username = generate_username(email)
    
    # Simpan ke database
    new_user = Users(
            username=username,
            password=generate_password_hash(secrets.token_urlsafe(12), method='pbkdf2:sha256'),
            nama_lengkap=user_info['name'],
            email=user_info['email'],
            jenis_kelamin=None,
            usia=None,
            foto=user_info['picture'],
            nomor_hp=None,
            level='peserta',
            reset_token=None,
            token_exp=None,
            login_method="google"
    )
    if Users.query.filter_by(email=email).first():
        flash("Email sudah digunakan. Silakan login.", "warning")
        return redirect(url_for('login'))
    if not new_user.jenis_kelamin:
        new_user.jenis_kelamin = "tidak_diketahui"
    db.session.add(new_user)
    db.session.commit()
    
    # Login langsung setelah registrasi
    login_user(new_user)
    session['username'] = new_user.username
    session['user'] = {
        'id': new_user.id,
        'username': new_user.username,
        'email': new_user.email,
        'nama_lengkap': new_user.nama_lengkap,
        'foto': new_user.foto,
        'level': new_user.level
    }
    flash("Registrasi berhasil! Selamat datang pengguna baru. Anda sekarang login untuk pertama kali.", "welcome")
    return redirect(url_for('admin_dashboard'))

# Endpoint Find Account
@app.route('/find_account/', methods=['GET', 'POST'])
def find_account():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        phone = request.form.get('no-hp')
        # Pastikan username diisi
        if not username:
            flash('Username wajib diisi.', 'danger')
            return redirect(url_for('find_account'))

        # ===== Kondisi 1: Username + Email =====
        if email and not phone:
            # Validasi format email
            email_pattern = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
            if not re.match(email_pattern, email):
                flash('Alamat email tidak valid.', 'danger')
                return redirect(url_for('find_account'))
            user_exists = check_username_in_db(username)
            email_exists = check_email_in_db(email)

            if user_exists and email_exists:
                verification_code = generate_verification_code()
                session['verification_code'] = verification_code
                session['verification_code_expiry'] = time.time() + 180
                session['username'] = username
                # Kirim kode via email
                msg = Message('Verifikasi Akun Anda',
                              sender='adip98816@gmail.com',
                              recipients=[email])
                safe_code = escape(verification_code)
                msg.html = f"""
                <p>Halo,</p>
                <p>Berikut adalah kode verifikasi 6 digit untuk mengakses akun Anda:</p>
                <h2>{safe_code}</h2>
                <p>Atau klik <a href="{url_for('verify_code', _external=True)}" style="color: blue;">link ini</a> untuk melanjutkan.</p>
                """
                mail.send(msg)
                flash(f'Kode verifikasi telah dikirim ke email {escape(email)}', 'success')
                return redirect(url_for('verify_code'))
            elif not user_exists and not email_exists:
                flash('Username dan email tidak ditemukan.', 'danger')
            elif not user_exists:
                flash('Username tidak ditemukan.', 'danger')
            elif not email_exists:
                flash('Email tidak ditemukan.', 'danger')
    
        # ===== Kondisi 2: Username + Nomor HP =====
        elif phone and not email:
            normalized_phone = normalize_phone_number(phone)
            phone_pattern = r'^\+628\d{7,12}$'
            if not re.match(phone_pattern, normalized_phone):
                flash('Format nomor HP tidak valid. Gunakan nomor Indonesia.', 'danger')
                return redirect(url_for('find_account'))
            # Cek keberadaan username dan nomor HP di database
            user_exists = check_username_in_db(username)
            hp_exists = check_phone_in_db(normalized_phone)
            if user_exists and hp_exists:
                verification_code = generate_verification_code()
                session['verification_code'] = verification_code
                session['verification_code_expiry'] = time.time() + 180  # berlaku 3 menit
                session['username'] = username
                session['phone'] = normalized_phone
                send_whatsapp_code(normalized_phone, verification_code)
                flash(f'Kode verifikasi telah dikirim ke nomor WhatsApp {normalized_phone}', 'success')
                return redirect(url_for('verify_code'))
            # Penanganan error spesifik
            if not user_exists and not hp_exists:
                flash('Username dan nomor HP tidak ditemukan.', 'danger')
            elif not user_exists:
                flash('Username tidak ditemukan.', 'danger')
            elif not hp_exists:
                flash('Nomor HP tidak ditemukan.', 'danger')
            return redirect(url_for('find_account'))
        else:
            flash('Harap isi email atau nomor HP.', 'danger')
        return redirect(url_for('find_account'))
    return render_template('find_account.html')

# Endpoint untuk verify_code
@app.route('/verify_code/', methods=['GET', 'POST'])
def verify_code():
    if request.method == 'GET':
        expiry_time = session.get('verification_code_expiry', 0)
        return render_template('verify_code.html', expiry_time=int(expiry_time))
    # Metode untuk memproses data JSON
    expiry_time = session.get('verification_code_expiry', 0)
    data = request.get_json()
    if not data or 'verification_code' not in data:
        return jsonify({'message': 'Verification code is required.'}), 400
    code = data['verification_code']
    if time.time() > expiry_time:
        return jsonify({'message': 'Verification code has expired.'}), 400
    if 'verification_code' in session and session['verification_code'] == int(code):
        # Generate reset token dan set waktu kedaluwarsa
        reset_token = secrets.token_hex(16)
        expiry_time = datetime.now() + timedelta(minutes=10)
        username = session.get('username')
        user = Users.query.filter_by(username=username).first()
        if not user:
            return jsonify({'message': 'User not found.'}), 404
        # Simpan token dan waktu kedaluwarsa di database lalu sertakan URL reset password dengan token
        user.reset_token = reset_token
        user.token_exp = expiry_time
        db.session.commit()
        reset_password_url = escape(url_for('reset_password', token=reset_token, _external=True))
        return jsonify({
            'message': 'Verification successful.',
            'redirect_url': reset_password_url
        }), 200
    else:
        return jsonify({'message': 'Incorrect verification code.'}), 400
    
# Endpoint Reset Password
@app.route('/reset_password/', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'GET':
        reset_token = request.args.get('token')
        if not reset_token:
            return jsonify({'error': "Token tidak ditemukan."}), 400
        # Validasi token
        user = Users.query.filter_by(reset_token=reset_token).first()
        if not user or datetime.now() > user.token_exp:
            if user:
                user.reset_token = None
                user.token_exp = None
                db.session.commit()
            return jsonify({'error': "Token tidak valid atau telah kedaluwarsa."}), 400
        # Jika token valid, arahkan ke halaman reset password
        return render_template('reset_password.html', token=escape(reset_token))
    if request.method == 'POST':
        data = request.get_json()
        reset_token = data.get('reset_token')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        # Debug input
        print(f"Reset token: {reset_token}, Password baru: {new_password}")
        # Validasi input
        if not reset_token or not new_password or not confirm_password:
            return jsonify({'error': "Semua data harus diisi."}), 400
        # Validasi pola password
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, new_password):
            return jsonify({'error': "Password harus terdiri dari minimal 8 karakter, termasuk huruf besar, kecil, angka, dan simbol."}), 400
        if new_password != confirm_password:
            return jsonify({'error': "Password dan konfirmasi password tidak cocok."}), 400
        # Validasi token
        user = Users.query.filter_by(reset_token=reset_token).first()
        if not user:
            return jsonify({'error': "Token reset password tidak valid."}), 400
        if datetime.now() > user.token_exp:
            user.reset_token = None
            user.token_exp = None
            db.session.commit()
            return jsonify({'error': "Token reset password telah kedaluwarsa."}), 400
        # Update password
        hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256', salt_length=16)
        user.password = hashed_password
        user.reset_token = None
        user.token_exp = None
        db.session.commit()
        print("Password berhasil diubah.")
        return jsonify({'message': "Password Anda telah berhasil diubah!"}), 200
    
# Route Index
@app.route("/")
@app.route("/index/")
def index():
    username = None
    user_data = None
    notification_count = 0
    profile_picture = None
    
    print("session keys:", session.keys())
    print("first_time_login:", session.get("first_time_login"))
    print("Session di /index/:", dict(session))
    first_time = session.get('first_time_login', None)

    # Cek login via session username (manual)
    if 'username' in session:
        username = session['username']
        user = Users.query.filter_by(username=username).first()
        if user:
            notification_count = Notification.query.filter_by(user_id=user.id, is_read=False).count()
            profile_picture = user.foto
        else:
            session.pop('username', None)

    # Cek login via session user (Google OAuth)
    elif 'user' in session:
        user_data = session['user'] 
        username = user_data.get('username')
        profile_picture = user_data.get('foto')
        user_id = user_data.get('id')
        if user_id:
            notification_count = Notification.query.filter_by(user_id=user_id, is_read=False).count()
    return render_template('index.html', username=username, profile_picture=profile_picture, notification_count=notification_count, user_data=user_data, first_time_login=first_time, debug_theme=session.get("theme"))

@app.route("/clear-first-login-flag", methods=["POST"])
@csrf.exempt
def clear_first_login_flag():
    session.pop('first_time_login', None)
    return '', 204

@app.before_request
def check_session():
    print("Session sekarang:", dict(session))
    
@app.route('/set-language/<lang_code>')
def set_language(lang_code):
    if lang_code in ['id', 'en']:
        session['lang'] = lang_code
    return redirect(request.referrer or url_for('index'))

@app.context_processor
def inject_current_lang():
    return dict(current_lang=session.get('lang', 'id'))

@app.route('/set-theme/<theme>', methods=['POST'])
def set_theme(theme):
    if theme in ['light', 'dark']:
        session['theme'] = theme
        session.modified = True
        return '', 204
    return 'Invalid theme', 400

@app.context_processor
def inject_theme():
    return dict(current_theme=session.get('theme', 'light'))

@app.route('/save_sidebar_state', methods=['POST'])
@login_required
def save_sidebar_state():
    data = request.get_json()
    state = data.get('state')

    if state not in ['expanded', 'collapsed']:
        return jsonify({'status': 'error', 'message': 'Invalid state'}), 400

    current_user.sidebar_state = state
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Sidebar state saved'})

# --- Route Dashboard Admin ---
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    sidebar_state = current_user.sidebar_state or 'expanded'
    if 'username' not in session:
        flash("Silakan login terlebih dahulu", "warning")
        return redirect(url_for('login'))
    
    user = current_user
    if not user:
        flash("Akses ditolak. User tidak valid!", "danger")
        return redirect(url_for('index'))
    
    # Cek level user
    if user.level == 'penilai':
        return redirect(url_for('penilai_dashboard'))
    elif user.level == 'peserta':
        return redirect(url_for('peserta_dashboard'))
    elif user.level != 'admin':
        flash("Akses ditolak. Anda bukan admin!", "danger")
        return redirect(url_for('index'))
    
    total_users = Users.query.count()
    total_participants = Participants.query.count() if db.inspect(db.engine).has_table("participants") else 0
    total_criteria = Criteria.query.count() if db.inspect(db.engine).has_table("criteria") else 0
    total_notifications = Notification.query.count()

    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('dashboard_admin.html', sidebar_state=sidebar_state, user=user, total_users=total_users, total_participants=total_participants, total_criteria=total_criteria, total_notifications=total_notifications, time=time)

# Middleware untuk membatasi akses hanya admin
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'role' not in session or session['role'] != 'admin':
            flash("Akses ditolak! Hanya admin yang bisa membuka halaman ini.", "error")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.all()
    users_data = [u.to_dict() for u in users]
    return render_template('manajemen_pengguna.html', sidebar_state=sidebar_state, users=users_data, time=time)

@app.route('/admin/add_user', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_user():
    sidebar_state = current_user.sidebar_state or 'expanded'

    if request.method == 'POST':
        # Ambil data dari form
        nama_lengkap = request.form.get('fullName')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form['confirmPassword']
        level = request.form.get('level')  

        # Validasi password dengan regex
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, password):
            flash("Password must have at least 8 characters, including uppercase, lowercase, number, and special character.", "danger")
            return redirect(url_for('admin_users'))
        # Validasi apakah password dan confirmPassword cocok
        if password != confirm_password:
            flash("Password and Confirm Password must match!", "danger")
            return redirect(url_for('admin_users'))
        # Cek keberadaan username dan email di database
        user_exists = check_username_in_db(username)
        email_exists = check_email_in_db(email)
        if not user_exists and not email_exists:
            # Enkripsi password
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16) 
            # Masukkan data ke database
            try:
                new_user = Users(
                    username=username,
                    password=hashed_password,
                    nama_lengkap=nama_lengkap,
                    email=email,
                    level=level,
                    status='aktif'
                )
                db.session.add(new_user)
                db.session.commit()
                flash("Akun berhasil dibuat!", "success")
                return redirect(url_for('admin_users'))
            except Exception as e:
                db.session.rollback()
                logging.error(f"Error during registration: {e}")
                flash("An error occurred during registration, Please try again.", "danger")
                print(e)
        elif user_exists and email_exists:
            flash("Username dan email Anda telah terdaftar.", "danger")
        elif user_exists:
            flash("Username Anda telah terdaftar.", "danger")
        elif email_exists:
            flash("Email Anda telah terdaftar.", "danger")
        return redirect(url_for('admin_users'))
    return render_template('manajemen_pengguna.html', sidebar_state=sidebar_state, users=Users.query.all(), time=time)

# Admin/Import Data User
def allowed_file(filename):
    allowed_extensions = {'csv', 'xls', 'xlsx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/admin/import_users', methods=['POST'])
@login_required
@admin_required
def admin_import_users():
    if 'file' not in request.files:
        flash('Tidak ada file yang diupload!', 'error')
        return redirect(url_for('admin_users'))
    
    file = request.files['file']
    if not file.filename:
        flash('Nama file tidak valid!', 'error')
        return redirect(url_for('admin_users'))

    if file.mimetype not in [
        'text/csv',
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]:
        flash('Tipe file tidak didukung!', 'error')
        return redirect(url_for('admin_users'))

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if not allowed_file(file.filename):
        flash('Format file tidak diizinkan! Gunakan CSV atau Excel.', 'error')
        return redirect(url_for('admin_users'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        df = pd.read_csv(filepath) if ext == 'csv' else pd.read_excel(filepath)
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_").str.replace("-", "_")

        required_cols = ['nama_lengkap', 'username', 'email', 'level']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            flash(f"Kolom berikut tidak ditemukan: {', '.join(missing)}", 'error')
            return redirect(url_for('admin_users'))

        existing_usernames = {u[0] for u in db.session.query(Users.username).all()}
        valid_levels = {'admin', 'penilai', 'peserta'}
        count_added, count_skipped = 0, 0
        new_users = []

        for _, row in df.iterrows():
            # Validasi
            if pd.isna(row['username']) or pd.isna(row['email']):
                count_skipped += 1
                continue
            if row['username'] in existing_usernames:
                count_skipped += 1
                continue
            if row['level'] not in valid_levels:
                count_skipped += 1
                continue
            if '@' not in str(row['email']):
                count_skipped += 1
                continue

            password = row['password'] if 'password' in df.columns and pd.notna(row['password']) else '12345678'
            new_users.append(Users(
                nama_lengkap=row['nama_lengkap'],
                username=row['username'],
                email=row['email'],
                password=generate_password_hash(str(password)),
                level=row['level'],
                jenis_kelamin=row.get('jenis_kelamin'),
                usia=row.get('usia', 0),
                nomor_hp=row.get('nomor_hp', "")
            ))
            count_added += 1

        if new_users:
            db.session.bulk_save_objects(new_users)
            db.session.commit()
            flash({
                'category': 'success',
                'title': 'Import Berhasil ✅',
                'message': f'{count_added} pengguna baru ditambahkan, {count_skipped} dilewati.'
            })
        else:
            flash({
                'category': 'danger',
                'title': 'Tidak Ada Data Baru ⚠️',
                'message': 'File sudah diproses, tetapi tidak ada pengguna baru yang ditambahkan.'
            })
    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Import gagal: {e}")
        flash(f'Terjadi kesalahan saat import: {e}', 'error')
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
    return redirect(url_for('admin_users'))

# Download Data User
@app.route('/download_users')
def download_users():
    users = Users.query.order_by(Users.nama_lengkap.asc()).all()
    
    # Jika tidak ada data pengguna
    if not users:
        flash({
            'category': 'warning',
            'title': 'Tidak Ada Data ⚠️',
            'message': 'Tidak ada data pengguna yang tersedia untuk diunduh.'
        })
        return redirect(url_for('admin_users'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pengguna"
    headers = ['No', 'Nama Lengkap', 'Username', 'Email', 'Level', 'Status']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for i, u in enumerate(users, start=1):
        ws.append([
            i,
            u.nama_lengkap or '',
            u.username or '',
            u.email or '',
            u.level or '',
            u.status or ''
        ])
        
    # Auto width untuk setiap kolom
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"data_pengguna_{timestamp}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

# Admin/Delete User
@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    # Hanya admin boleh hapus
    if current_user.level != "admin":
        flash("Anda tidak memiliki izin untuk menghapus pengguna.", "danger")
        return redirect(url_for('admin_users'))
    
    user = Users.query.get(user_id)
    if not user:
        flash("Pengguna tidak ditemukan!", "danger")
        return redirect(url_for('admin_users'))
    
    # Proteksi: admin tidak bisa menghapus dirinya sendiri
    if user.id == current_user.id:
        flash("Anda tidak dapat menghapus akun Anda sendiri!", "warning")
        return redirect(url_for('admin_users'))
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f"Pengguna '{user.username}' berhasil dihapus!", "success")
        logging.info(f"User '{user.username}' berhasil dihapus oleh admin.")
    except Exception as e:
        db.session.rollback()
        flash("Terjadi kesalahan saat menghapus data.", "danger")
        logging.error(f"Gagal menghapus user_id {user_id}: {e}")
    return redirect(url_for('admin_users'))

# Admin/Edit User
@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = Users.query.get(user_id)
    if not user:
        flash("Pengguna tidak ditemukan!", "danger")
        return redirect(url_for('admin_users'))
    
    user.nama_lengkap = request.form.get('nama_lengkap')
    user.email = request.form.get('email')
    user.username = request.form.get('username')
    user.level = request.form.get('level')
    user.status = request.form.get('status')
    user.jenis_kelamin = request.form.get('jenis_kelamin')
    user.usia = request.form.get('usia') or user.usia
    user.nomor_hp = request.form.get('nomor_hp') or user.nomor_hp

    # kalau ada file foto
    foto_file = request.files.get('foto')
    if foto_file and foto_file.filename:
        # proses simpan file sesuai implementasimu
        # simpan path ke user.foto = 'img/xxx.png'
        pass

    try:
        db.session.commit()
        flash(f"Data pengguna '{user.username}' berhasil diperbarui!", "success")
    except Exception as e:
        db.session.rollback()
        flash("Terjadi kesalahan saat memperbarui data.", "danger")
        logging.error(f"Gagal memperbarui user_id {user_id}: {e}")
    return redirect(url_for('admin_users'))

# Manajemen Seleksi   
@app.route('/admin/manajemen_seleksi')
@login_required
@admin_required
def admin_manajemen_seleksi():
    sidebar_state = current_user.sidebar_state or 'expanded'
    kegiatan_list = Event.query.all()
    return render_template("manajemen_seleksi.html", kegiatan_list=kegiatan_list, sidebar_state=sidebar_state, user=current_user, time=time)

# Konfigurasi Seleksi
@app.route('/api/save_config', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def save_config():
    try:
        data = request.get_json(force=True)
        activities = data.get('activities', [])
        criteria_list = data.get('criteria', [])

        if not activities and not criteria_list:
            return jsonify({'status': 'error', 'message': 'No data provided'}), 400
        created_events = []

        # Buat Event & Kuota
        for act in activities:
            nama = (act.get('nama') or '').strip()
            if not nama:
                continue
            mulai_date = None
            selesai_date = None
            try:
                if act.get('mulai'):
                    mulai_date = datetime.strptime(act['mulai'], '%Y-%m-%d').date()
                if act.get('selesai'):
                    selesai_date = datetime.strptime(act['selesai'], '%Y-%m-%d').date()
            except Exception:
                pass
            event = Event(
                jenis_kegiatan=act.get('jenis', 'siaga'),
                nama_kegiatan=nama,
                waktu_pelaksanaan=act.get('waktu', ''),
                tempat_pelaksanaan=act.get('tempat', act.get('tempat', '-')),
                skala_kegiatan=act.get('skala', 'ranting'),
                kwartir_penyelenggara=act.get('kwartir', 'kwartir ranting'),
                mulai=mulai_date or datetime.utcnow().date(),
                selesai=selesai_date or (mulai_date or datetime.utcnow().date())
            )
            db.session.add(event)
            db.session.flush()
            kuota = Kuota(
                event_id=event.id_kegiatan,
                putra=int(act.get('putra') or 0),
                putri=int(act.get('putri') or 0)
            )
            db.session.add(kuota)
            created_events.append(event)
        target_event_id = created_events[0].id_kegiatan if created_events else None

        # Buat Criteria
        for c in criteria_list:
            nama_kriteria = (c.get('nama') or '').strip() or 'Unnamed Criteria'
            bobot = float(c.get('bobot') or 0)
            aspek = c.get('aspek', [])
            aspek_str = ', '.join(aspek) if isinstance(aspek, list) else (aspek or '')
            jumlah_soal = c.get('jumlah_soal') or c.get('jumlahSoal') or None
            deskripsi = c.get('deskripsi', '')
            jenis_kriteria = c.get('jenis_kriteria', 'umum')
            if target_event_id is None:
                placeholder = Event(
                    jenis_kegiatan='-',
                    nama_kegiatan='(Default) Konfigurasi Seleksi',
                    waktu_pelaksanaan='',
                    tempat_pelaksanaan='-',
                    skala_kegiatan='-',
                    kwartir_penyelenggara='-',
                    mulai=datetime.utcnow().date(),
                    selesai=datetime.utcnow().date()
                )
                db.session.add(placeholder)
                db.session.flush()
                target_event_id = placeholder.id_kegiatan
            crit = Criteria(
                event_id=target_event_id,
                nama_kriteria=nama_kriteria,
                aspek=aspek_str,
                bobot=bobot,
                deskripsi=deskripsi,
                jenis_kriteria=jenis_kriteria,
                jumlah_soal=int(jumlah_soal) if jumlah_soal else None
            )
            db.session.add(crit)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Konfigurasi berhasil disimpan'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/save_config:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Kegiatan 
@app.route('/api/kegiatan')
@login_required
@admin_required
def api_kegiatan():
    kegiatan = Event.query.all()
    result = [
        {
            "id": k.id_kegiatan,
            "nama_kegiatan": k.nama_kegiatan,
            "jenis_kegiatan": k.jenis_kegiatan,
            "waktu_pelaksanaan": k.waktu_pelaksanaan.strftime("%Y-%m-%d"),
            "tempat_pelaksanaan": k.tempat_pelaksanaan,
            "skala_kegiatan": k.skala_kegiatan,
            "kwartir_penyelenggara": k.kwartir_penyelenggara,
            "mulai": k.mulai.strftime("%Y-%m-%d"),
            "selesai": k.selesai.strftime("%Y-%m-%d"),
        }
        for k in kegiatan
    ]
    return jsonify(result)

# API Kuota Kegiatan
@app.route('/api/kuota/<int:event_id>')
@login_required
@admin_required
def api_kuota(event_id):
    event = Event.query.get_or_404(event_id)
    if not event.kuota:
        return jsonify({"putra": 0, "putri": 0})

    # asumsi tabel Kuota punya kolom putra & putri
    kuota = event.kuota[0]
    return jsonify({
        "putra": kuota.putra,
        "putri": kuota.putri
    })

# API Data Peserta
@app.route("/api/peserta/<int:kegiatan_id>")
def get_peserta(kegiatan_id):
    peserta = Participants.query.filter_by(kegiatan_id=kegiatan_id).all()
    data = []
    for p in peserta:
        data.append({
            "nama_lengkap": p.nama_lengkap,
            "tanggal_lahir": str(p.tanggal_lahir),
            "jenis_kelamin": p.jenis_kelamin,
            "usia": p.usia,
            "alamat_tinggal": p.alamat_tinggal,
            "golongan": p.golongan,
            "tingkatan": p.tingkatan,
            "asal_gudep": p.asal_gudep,
            "asal_kwarran": p.asal_kwarran,
            "asal_kwarcab": p.asal_kwarcab,
            "asal_kwarda": p.asal_kwarda,
            "nomor_hp": p.nomor_hp,
            "email": p.email
        })
    return jsonify(data)

# Tambah Kegiatan
@app.route('/admin/tambah_seleksi', methods=['GET', 'POST'])
@login_required
@admin_required
def tambah_seleksi():
    if request.method == 'POST':
        nama = request.form['nama_kegiatan']
        jenis = request.form['jenis_kegiatan']
        waktu = request.form['waktu_pelaksanaan']
        tempat = request.form['tempat_pelaksanaan']
        skala = request.form['skala_kegiatan']
        kwartir = request.form['kwartir_penyelenggara']

        new_event = Event(
            nama_kegiatan=nama,
            jenis_kegiatan=jenis,
            waktu_pelaksanaan=waktu,
            tempat_pelaksanaan=tempat,
            skala_kegiatan=skala,
            kwartir_penyelenggara=kwartir
        )
        db.session.add(new_event)
        db.session.commit()

        flash('Kegiatan berhasil ditambahkan!', 'success')
        return redirect(url_for('admin_manajemen_seleksi'))
    return render_template("tambah_kegiatan.html")

# Edit Kegiatan
@app.route('/admin/edit_kegiatan/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_kegiatan(id):
    event = Event.query.get_or_404(id)
    if request.method == 'POST':
        event.nama_kegiatan = request.form['nama_kegiatan']
        event.jenis_kegiatan = request.form['jenis_kegiatan']
        event.waktu_pelaksanaan = request.form['waktu_pelaksanaan']
        event.tempat_pelaksanaan = request.form['tempat_pelaksanaan']
        db.session.commit()
        flash('Kegiatan berhasil diupdate!', 'success')
        return redirect(url_for('admin_manajemen_seleksi'))
    return render_template("edit_kegiatan.html", event=event)

# Hapus Kegiatan
@app.route('/admin/hapus_kegiatan/<int:id>', methods=['GET'])
@login_required
@admin_required
def hapus_kegiatan(id):
    event = Event.query.get_or_404(id)
    db.session.delete(event)
    db.session.commit()
    flash('Kegiatan berhasil dihapus!', 'danger')
    return redirect(url_for('admin_manajemen_seleksi'))

@app.route('/admin/detail_kegiatan/<int:id>')
@login_required
@admin_required
def detail_kegiatan(id):
    event = Event.query.get_or_404(id)
    return render_template("detail_kegiatan.html", event=event)

@app.route('/admin/kriteria')
@login_required
@admin_required
def admin_kriteria():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('data_kriteria.html', sidebar_state=sidebar_state, user=users, time=time)
    
@app.route('/admin/pembobotan_kriteria')
@login_required
@admin_required
def admin_pembobotan_kriteria():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('pembobotan_kriteria.html', sidebar_state=sidebar_state, user=users, time=time)

@app.route('/admin/peserta')
@login_required
@admin_required
def admin_peserta():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('data_peserta.html', sidebar_state=sidebar_state, user=users, time=time)
    
@app.route('/admin/hasil_seleksi')
@login_required
@admin_required
def admin_hasil_seleksi():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('hasil_seleksi.html', sidebar_state=sidebar_state, user=users, time=time)

@app.route('/admin/notifikasi')
@login_required
@admin_required
def admin_notifikasi():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('notifikasi.html', sidebar_state=sidebar_state, user=users, time=time)
    
@app.route('/admin/log_aktivitas')
@login_required
@admin_required
def admin_log_aktivitas():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('log_aktivity.html', sidebar_state=sidebar_state, user=users, time=time)

@app.route('/admin/settings')
@login_required
@admin_required
def admin_settings():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('settings.html', sidebar_state=sidebar_state, user=users, time=time)
    
@app.route('/penilai/dashboard')
@login_required
def penilai_dashboard():
    if current_user.level != 'penilai':
        flash("Anda tidak memiliki akses ke halaman ini.", "error")
        return redirect(url_for('index'))

    # Contoh data yang bisa ditampilkan di dashboard penilai
    data_peserta = Participants.query.all()  # Ambil semua peserta
    total_peserta = Participants.query.count()
    total_penilai = Users.query.filter_by(level='penilai').count()

    return render_template(
        'penilai/dashboard.html',
        data_peserta=data_peserta,
        total_peserta=total_peserta,
        total_penilai=total_penilai
    )

@app.route('/peserta/dashboard')
@login_required
def peserta_dashboard():
    if current_user.level != 'peserta':
        flash("Anda tidak memiliki akses ke halaman ini.", "error")
        return redirect(url_for('index'))

    # Contoh data untuk peserta
    biodata = Participants.query.filter_by(user_id=current_user.id).first()
    status_seleksi = biodata.status if biodata else "Belum ada status"
    nilai_akhir = biodata.nilai if biodata else None

    return render_template(
        'peserta/dashboard.html',
        biodata=biodata,
        status_seleksi=status_seleksi,
        nilai_akhir=nilai_akhir
    )

@app.route('/logout/')
def logout():
    session.clear()
    session.pop('username', None)
    flash("Anda telah logout.", "info")
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
